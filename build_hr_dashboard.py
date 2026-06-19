from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.worksheet.datavalidation import DataValidation
import random

wb = Workbook()

# ─── COLOR PALETTE ───────────────────────────────────────────────────────────
DARK_BG   = "1E3A5F"   # dark navy
MID_BG    = "2E5C9E"   # medium blue
ACCENT    = "00B4D8"   # cyan accent
LIGHT_BG  = "EAF4FB"   # very light blue
WHITE     = "FFFFFF"
GOLD      = "F4A261"
RED       = "E63946"
GREEN     = "2DC653"
GREY_BG   = "F5F7FA"
BORDER_C  = "BFD3E6"
TEXT_DARK = "1A1A2E"
TEXT_MID  = "2E5C9E"

def thin_border(color=BORDER_C):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=10, bold=True, color=WHITE):
    return Font(name='Segoe UI', size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=TEXT_DARK):
    return Font(name='Segoe UI', size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def vcenter():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – RAW DATA (Employee Table)
# ════════════════════════════════════════════════════════════════════════════
ws_data = wb.active
ws_data.title = "Employee_Data"

depts  = ["Sales","Engineering","HR","Finance","Marketing","Operations","IT"]
statuses = ["Active","Active","Active","Active","On Leave","Resigned"]
genders  = ["Male","Female","Male","Female","Male"]
locations = ["Madurai","Chennai","Bangalore","Hyderabad","Mumbai","Delhi","Pune"]
ratings   = ["Excellent","Good","Average","Below Average","Excellent","Good"]

random.seed(42)

headers = ["Emp_ID","Name","Department","Gender","Location","Join_Date",
           "Age","Salary","Performance_Rating","Status","Years_Exp",
           "Training_Hours","Manager"]

managers = {
    "Sales":"Ravi Kumar","Engineering":"Priya Nair","HR":"Sunita Sharma",
    "Finance":"Arun Mehta","Marketing":"Divya Rao","Operations":"Karthik Iyer",
    "IT":"Meena Pillai"
}

col_widths = [10,18,14,9,12,13,7,12,18,10,12,16,16]

for i, h in enumerate(headers, 1):
    c = ws_data.cell(row=1, column=i, value=h)
    c.font  = header_font(10)
    c.fill  = fill(DARK_BG)
    c.alignment = center()
    ws_data.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

ws_data.row_dimensions[1].height = 28

first_names = ["Aarav","Vivaan","Ananya","Diya","Rohan","Priya","Arjun",
               "Shreya","Kiran","Neha","Raj","Sita","Mohan","Kavya","Dev",
               "Nisha","Arun","Pooja","Vijay","Meera","Suresh","Lakshmi",
               "Vikram","Deepa","Ravi","Anjali","Sanjay","Sunita","Rahul",
               "Divya","Manoj","Rekha","Nikhil","Swati","Aditya","Prathyusha",
               "Harish","Yamini","Ganesh","Saranya","Bala","Lavanya","Senthil",
               "Gomathi","Murugan","Vijayalakshmi","Prakash","Revathi","Dinesh","Malathi"]

last_names = ["Sharma","Patel","Iyer","Nair","Kumar","Singh","Reddy","Mehta",
              "Pillai","Rao","Gupta","Joshi","Verma","Menon","Subramanian"]

rows = []
for i in range(1, 101):
    emp_id   = f"EMP{i:03d}"
    name     = f"{random.choice(first_names)} {random.choice(last_names)}"
    dept     = random.choice(depts)
    gender   = random.choice(genders)
    loc      = random.choice(locations)
    yr       = random.randint(2016, 2023)
    mo       = random.randint(1, 12)
    dy       = random.randint(1, 28)
    join     = f"{dy:02d}/{mo:02d}/{yr}"
    age      = random.randint(22, 52)
    base     = {"Sales":35000,"Engineering":55000,"HR":32000,"Finance":45000,
                "Marketing":38000,"Operations":30000,"IT":52000}
    salary   = round(base[dept] * random.uniform(0.85, 1.60), -2)
    rating   = random.choice(ratings)
    status   = random.choice(statuses)
    exp      = random.randint(1, 20)
    train    = random.randint(10, 80)
    mgr      = managers[dept]
    rows.append([emp_id,name,dept,gender,loc,join,age,salary,rating,status,exp,train,mgr])

for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        cell = ws_data.cell(row=r, column=c, value=val)
        cell.font      = body_font(9)
        cell.alignment = vcenter()
        cell.border    = thin_border()
        bg = WHITE if r % 2 == 0 else GREY_BG
        cell.fill = fill(bg)
        if c == 8:  # Salary
            cell.number_format = '₹#,##0'
        if c == 10:  # Status colour
            if val == "Resigned":
                cell.font = body_font(9, color=RED)
            elif val == "On Leave":
                cell.font = body_font(9, color=GOLD)
            else:
                cell.font = body_font(9, color=GREEN)

ws_data.freeze_panes = "A2"

# Conditional formatting – salary heat
ws_data.conditional_formatting.add(
    "H2:H101",
    ColorScaleRule(start_type='min', start_color='E8F5E9',
                   end_type='max',   end_color='1B5E20')
)
# Performance data bar
ws_data.conditional_formatting.add(
    "L2:L101",
    DataBarRule(start_type='min', start_value=0,
                end_type='max', end_value=None,
                color=ACCENT)
)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – SUMMARY (Pivot-style formulas)
# ════════════════════════════════════════════════════════════════════════════
ws_sum = wb.create_sheet("Summary")

def section_header(ws, row, col, title, span=4):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=title)
    c.font = header_font(11)
    c.fill = fill(MID_BG)
    c.alignment = center()

def kpi_block(ws, row, col, label, value_formula, fmt='#,##0', color=DARK_BG):
    ws.merge_cells(start_row=row,   start_column=col,
                   end_row=row,     end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col,
                   end_row=row+1,   end_column=col+1)
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = header_font(9, color=WHITE)
    lc.fill = fill(color)
    lc.alignment = center()
    vc = ws.cell(row=row+1, column=col, value=value_formula)
    vc.font = Font(name='Segoe UI', size=14, bold=True, color=color)
    vc.fill = fill(LIGHT_BG)
    vc.alignment = center()
    vc.number_format = fmt
    ws.row_dimensions[row].height   = 22
    ws.row_dimensions[row+1].height = 32

# ── Top banner ──────────────────────────────────────────────────────────────
ws_sum.merge_cells("A1:N1")
banner = ws_sum.cell(row=1, column=1,
    value="🏢  HR ANALYTICS DASHBOARD  |  Workforce Intelligence Report")
banner.font  = Font(name='Segoe UI', size=16, bold=True, color=WHITE)
banner.fill  = fill(DARK_BG)
banner.alignment = center()
ws_sum.row_dimensions[1].height = 40

ws_sum.merge_cells("A2:N2")
sub = ws_sum.cell(row=2, column=1,
    value='Powered by COUNTIFS · AVERAGEIFS · SUMIFS · Dynamic Excel Formulas')
sub.font = Font(name='Segoe UI', size=9, italic=True, color=MID_BG)
sub.fill = fill(LIGHT_BG)
sub.alignment = center()

# ── KPI Row ─────────────────────────────────────────────────────────────────
ws_sum.row_dimensions[4].height = 14
section_header(ws_sum, 4, 1, "📊  KEY METRICS AT A GLANCE", span=14)

kpi_block(ws_sum, 5, 1,  "Total Employees",
          "=COUNTA(Employee_Data!A2:A1000)", "#,##0", DARK_BG)
kpi_block(ws_sum, 5, 3,  "Active Employees",
          '=COUNTIF(Employee_Data!J2:J1000,"Active")', "#,##0", MID_BG)
kpi_block(ws_sum, 5, 5,  "Avg Monthly Salary (₹)",
          "=AVERAGE(Employee_Data!H2:H1000)", "₹#,##0", "1A6B4A")
kpi_block(ws_sum, 5, 7,  "Total Payroll (₹)",
          "=SUM(Employee_Data!H2:H1000)", "₹#,##0", "7B2D8B")
kpi_block(ws_sum, 5, 9,  "Avg Age",
          "=ROUND(AVERAGE(Employee_Data!G2:G1000),1)", "#,##0.0", "C0392B")
kpi_block(ws_sum, 5, 11, "Avg Exp (Yrs)",
          "=ROUND(AVERAGE(Employee_Data!K2:K1000),1)", "#,##0.0", "D35400")
kpi_block(ws_sum, 5, 13, "Attrition Rate %",
          '=COUNTIF(Employee_Data!J2:J1000,"Resigned")/COUNTA(Employee_Data!A2:A1000)',
          "0.00%", RED)

# ── Dept Breakdown ───────────────────────────────────────────────────────────
ws_sum.row_dimensions[9].height = 14
section_header(ws_sum, 9, 1, "📁  DEPARTMENT BREAKDOWN", span=8)

dept_h = ["Department","Headcount","Active","Resigned","Avg Salary (₹)",
          "Avg Exp","Avg Training Hrs","% of Workforce"]
for i, h in enumerate(dept_h, 1):
    c = ws_sum.cell(row=10, column=i, value=h)
    c.font = header_font(9)
    c.fill = fill(MID_BG)
    c.alignment = center()

dept_list = ["Sales","Engineering","HR","Finance","Marketing","Operations","IT"]
for r, dept in enumerate(dept_list, 11):
    ws_sum.cell(row=r, column=1, value=dept).font = body_font(9, bold=True)
    ws_sum.cell(row=r, column=2,
        value=f'=COUNTIF(Employee_Data!C$2:C$1000,A{r})').number_format = "#,##0"
    ws_sum.cell(row=r, column=3,
        value=f'=COUNTIFS(Employee_Data!C$2:C$1000,A{r},Employee_Data!J$2:J$1000,"Active")').number_format = "#,##0"
    ws_sum.cell(row=r, column=4,
        value=f'=COUNTIFS(Employee_Data!C$2:C$1000,A{r},Employee_Data!J$2:J$1000,"Resigned")').number_format = "#,##0"
    ws_sum.cell(row=r, column=5,
        value=f'=IFERROR(AVERAGEIF(Employee_Data!C$2:C$1000,A{r},Employee_Data!H$2:H$1000),0)').number_format = "₹#,##0"
    ws_sum.cell(row=r, column=6,
        value=f'=IFERROR(AVERAGEIF(Employee_Data!C$2:C$1000,A{r},Employee_Data!K$2:K$1000),0)').number_format = "#,##0.0"
    ws_sum.cell(row=r, column=7,
        value=f'=IFERROR(AVERAGEIF(Employee_Data!C$2:C$1000,A{r},Employee_Data!L$2:L$1000),0)').number_format = "#,##0.0"
    ws_sum.cell(row=r, column=8,
        value=f'=IFERROR(B{r}/SUM($B$11:$B$17),0)').number_format = "0.0%"
    for c in range(1, 9):
        cell = ws_sum.cell(row=r, column=c)
        cell.border    = thin_border()
        cell.alignment = center()
        bg = WHITE if r % 2 == 0 else LIGHT_BG
        if not cell.fill or cell.fill.patternType == 'none':
            cell.fill = fill(bg)
        if c != 1:
            cell.font = body_font(9)

# Totals row
tr = 18
ws_sum.cell(row=tr, column=1, value="TOTAL").font = header_font(9, color=WHITE)
ws_sum.cell(row=tr, column=1).fill = fill(DARK_BG)
ws_sum.cell(row=tr, column=2, value="=SUM(B11:B17)").number_format = "#,##0"
ws_sum.cell(row=tr, column=5, value="=AVERAGE(E11:E17)").number_format = "₹#,##0"
ws_sum.cell(row=tr, column=8, value="=SUM(H11:H17)").number_format = "0.0%"
for c in range(1, 9):
    ws_sum.cell(row=tr, column=c).border = thin_border()
    ws_sum.cell(row=tr, column=c).alignment = center()
    ws_sum.cell(row=tr, column=c).fill = fill(DARK_BG)
    ws_sum.cell(row=tr, column=c).font = header_font(9)

# ── Performance Distribution ─────────────────────────────────────────────────
section_header(ws_sum, 9, 10, "🏆  PERFORMANCE SPLIT", span=5)
perf_h = ["Rating","Count","% Share"]
for i, h in enumerate(perf_h, 10):
    c = ws_sum.cell(row=10, column=i, value=h)
    c.font = header_font(9)
    c.fill = fill(MID_BG)
    c.alignment = center()

perf_ratings = ["Excellent","Good","Average","Below Average"]
perf_colors  = [GREEN,"2196F3",GOLD,RED]
for ri, (rat, col) in enumerate(zip(perf_ratings, perf_colors), 11):
    ws_sum.cell(row=ri, column=10, value=rat).font = body_font(9, bold=True, color=col)
    cnt = ws_sum.cell(row=ri, column=11,
        value=f'=COUNTIF(Employee_Data!I$2:I$1000,J{ri})')
    cnt.number_format = "#,##0"
    pct = ws_sum.cell(row=ri, column=12,
        value=f'=IFERROR(K{ri}/SUM($K$11:$K$14),0)')
    pct.number_format = "0.0%"
    for c in range(10, 13):
        ws_sum.cell(row=ri, column=c).border    = thin_border()
        ws_sum.cell(row=ri, column=c).alignment = center()
        ws_sum.cell(row=ri, column=c).fill = fill(WHITE if ri%2==0 else LIGHT_BG)

# ── Gender Split ─────────────────────────────────────────────────────────────
section_header(ws_sum, 16, 10, "⚥  GENDER SPLIT", span=5)
for i, h in enumerate(["Gender","Count","% Share"], 10):
    c = ws_sum.cell(row=17, column=i, value=h)
    c.font = header_font(9); c.fill = fill(MID_BG); c.alignment = center()
for ri, gender in enumerate(["Male","Female"], 18):
    ws_sum.cell(row=ri, column=10, value=gender).font = body_font(9, bold=True)
    ws_sum.cell(row=ri, column=11,
        value=f'=COUNTIF(Employee_Data!D$2:D$1000,J{ri})').number_format="#,##0"
    ws_sum.cell(row=ri, column=12,
        value=f'=IFERROR(K{ri}/SUM($K$18:$K$19),0)').number_format="0.0%"
    for c in range(10, 13):
        ws_sum.cell(row=ri, column=c).border = thin_border()
        ws_sum.cell(row=ri, column=c).alignment = center()
        ws_sum.cell(row=ri, column=c).fill = fill(WHITE if ri%2==0 else LIGHT_BG)

# ── Conditional formatting on dept headcount ─────────────────────────────────
ws_sum.conditional_formatting.add(
    "B11:B17",
    DataBarRule(start_type='min', start_value=0,
                end_type='max', end_value=None, color=ACCENT)
)

# Column widths for Summary
widths = [16,11,9,10,15,9,16,12, 2, 16,9,10]
for i, w in enumerate(widths, 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – CHARTS
# ════════════════════════════════════════════════════════════════════════════
ws_chart = wb.create_sheet("Charts")

ws_chart.merge_cells("A1:P1")
ch_banner = ws_chart.cell(row=1, column=1, value="📈  VISUAL ANALYTICS  –  HR Dashboard")
ch_banner.font  = Font(name='Segoe UI', size=14, bold=True, color=WHITE)
ch_banner.fill  = fill(DARK_BG)
ch_banner.alignment = center()
ws_chart.row_dimensions[1].height = 36

# Bar Chart – Headcount by Dept
bar = BarChart()
bar.type  = "col"
bar.title = "Headcount by Department"
bar.y_axis.title = "Employees"
bar.style = 10
bar.width  = 14; bar.height = 10
cats = Reference(ws_sum, min_col=1, max_col=1, min_row=11, max_row=17)
data = Reference(ws_sum, min_col=2, max_col=2, min_row=10, max_row=17)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = ACCENT
ws_chart.add_chart(bar, "A3")

# Pie Chart – Performance Distribution
pie = PieChart()
pie.title  = "Performance Rating Distribution"
pie.style  = 10
pie.width  = 14; pie.height = 10
pie_data = Reference(ws_sum, min_col=11, max_col=11, min_row=10, max_row=14)
pie_cats = Reference(ws_sum, min_col=10, max_col=10, min_row=11, max_row=14)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie_colors = ["00B050","4472C4","FFC000","FF0000"]
for idx, color in enumerate(pie_colors):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    pie.series[0].data_points.append(pt)
ws_chart.add_chart(pie, "I3")

# Bar Chart – Avg Salary by Dept
sal = BarChart()
sal.type  = "bar"
sal.title = "Average Salary by Department (₹)"
sal.x_axis.title = "Avg Salary"
sal.style = 10
sal.width  = 14; sal.height = 10
sal_data = Reference(ws_sum, min_col=5, max_col=5, min_row=10, max_row=17)
sal.add_data(sal_data, titles_from_data=True)
sal.set_categories(cats)
sal.series[0].graphicalProperties.solidFill = GOLD
ws_chart.add_chart(sal, "A22")

# Line Chart – Training Hrs by Dept
line = LineChart()
line.title = "Avg Training Hours by Department"
line.y_axis.title = "Hours"
line.style = 10
line.width  = 14; line.height = 10
line_data = Reference(ws_sum, min_col=7, max_col=7, min_row=10, max_row=17)
line.add_data(line_data, titles_from_data=True)
line.set_categories(cats)
line.series[0].graphicalProperties.line.solidFill = MID_BG
ws_chart.add_chart(line, "I22")

for col in "ABCDEFGHIJKLMNOP":
    ws_chart.column_dimensions[col].width = 10

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 – ATTRITION ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
ws_attr = wb.create_sheet("Attrition_Analysis")

ws_attr.merge_cells("A1:L1")
ab = ws_attr.cell(row=1, column=1,
    value="📉  ATTRITION & RETENTION ANALYSIS")
ab.font  = Font(name='Segoe UI', size=14, bold=True, color=WHITE)
ab.fill  = fill(RED); ab.alignment = center()
ws_attr.row_dimensions[1].height = 36

section_header(ws_attr, 3, 1, "ATTRITION BY DEPARTMENT", span=5)
attr_heads = ["Department","Total","Resigned","On Leave","Attrition Rate %"]
for i, h in enumerate(attr_heads, 1):
    c = ws_attr.cell(row=4, column=i, value=h)
    c.font = header_font(9); c.fill = fill(MID_BG); c.alignment = center()

for ri, dept in enumerate(dept_list, 5):
    ws_attr.cell(row=ri, column=1, value=dept).font = body_font(9, bold=True)
    ws_attr.cell(row=ri, column=2,
        value=f'=COUNTIF(Employee_Data!C$2:C$1000,A{ri})').number_format="#,##0"
    ws_attr.cell(row=ri, column=3,
        value=f'=COUNTIFS(Employee_Data!C$2:C$1000,A{ri},Employee_Data!J$2:J$1000,"Resigned")').number_format="#,##0"
    ws_attr.cell(row=ri, column=4,
        value=f'=COUNTIFS(Employee_Data!C$2:C$1000,A{ri},Employee_Data!J$2:J$1000,"On Leave")').number_format="#,##0"
    rate = ws_attr.cell(row=ri, column=5,
        value=f'=IFERROR(C{ri}/B{ri},0)')
    rate.number_format = "0.0%"
    for c in range(1, 6):
        ws_attr.cell(row=ri, column=c).border    = thin_border()
        ws_attr.cell(row=ri, column=c).alignment = center()
        bg = WHITE if ri%2==0 else LIGHT_BG
        ws_attr.cell(row=ri, column=c).fill = fill(bg)

# Conditional format: red for high attrition
ws_attr.conditional_formatting.add(
    "E5:E11",
    ColorScaleRule(start_type='min',  start_color='C8E6C9',
                   mid_type='percent',mid_value=50,mid_color='FFF9C4',
                   end_type='max',    end_color='FFCDD2')
)

# Experience vs Attrition table
section_header(ws_attr, 3, 7, "EXP. BAND vs ATTRITION", span=4)
exp_h = ["Exp Band","Total","Resigned","Rate %"]
for i, h in enumerate(exp_h, 7):
    c = ws_attr.cell(row=4, column=i, value=h)
    c.font = header_font(9); c.fill = fill(MID_BG); c.alignment = center()

bands = [("0-2 Yrs","<=2"),("3-5 Yrs","<=5"),("6-10 Yrs","<=10"),("11+ Yrs",">10")]
prev = [0, 2, 5, 10]
for bi, ((label, cond), lo) in enumerate(zip(bands, prev), 5):
    ws_attr.cell(row=bi, column=7, value=label).font = body_font(9, bold=True)
    hi = cond.replace("<=","").replace(">","")
    if "<=" in cond:
        cnt_f   = f'=COUNTIFS(Employee_Data!K$2:K$1000,">{lo}",Employee_Data!K$2:K$1000,"{cond}")'
        res_f   = f'=COUNTIFS(Employee_Data!K$2:K$1000,">{lo}",Employee_Data!K$2:K$1000,"{cond}",Employee_Data!J$2:J$1000,"Resigned")'
    else:
        cnt_f   = f'=COUNTIF(Employee_Data!K$2:K$1000,"{cond}")'
        res_f   = f'=COUNTIFS(Employee_Data!K$2:K$1000,"{cond}",Employee_Data!J$2:J$1000,"Resigned")'
    ws_attr.cell(row=bi, column=8, value=cnt_f).number_format="#,##0"
    ws_attr.cell(row=bi, column=9, value=res_f).number_format="#,##0"
    ws_attr.cell(row=bi, column=10,
        value=f'=IFERROR(I{bi}/H{bi},0)').number_format="0.0%"
    for c in range(7, 11):
        ws_attr.cell(row=bi, column=c).border    = thin_border()
        ws_attr.cell(row=bi, column=c).alignment = center()
        ws_attr.cell(row=bi, column=c).fill = fill(WHITE if bi%2==0 else LIGHT_BG)
        ws_attr.cell(row=bi, column=c).font = body_font(9)

attr_widths = [16,9,10,10,14,2,16,9,10,10]
for i, w in enumerate(attr_widths, 1):
    ws_attr.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 – EMPLOYEE LOOKUP (XLOOKUP-style with INDEX-MATCH)
# ════════════════════════════════════════════════════════════════════════════
ws_look = wb.create_sheet("Employee_Lookup")

ws_look.merge_cells("A1:G1")
lb = ws_look.cell(row=1, column=1,
    value="🔍  EMPLOYEE LOOKUP TOOL  –  Enter Emp ID below")
lb.font  = Font(name='Segoe UI', size=13, bold=True, color=WHITE)
lb.fill  = fill(DARK_BG); lb.alignment = center()
ws_look.row_dimensions[1].height = 36

ws_look.cell(row=3, column=1, value="Enter Employee ID:").font = body_font(10, bold=True)
inp = ws_look.cell(row=3, column=2, value="EMP001")
inp.font   = Font(name='Segoe UI', size=11, bold=True, color=RED)
inp.fill   = fill("FFF9C4")
inp.border = thin_border(RED)
inp.alignment = center()

fields = ["Name","Department","Gender","Location","Join Date",
          "Age","Monthly Salary","Performance","Status",
          "Years Experience","Training Hours","Manager"]
src_cols = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]

ws_look.cell(row=5, column=1,
    value="─── Employee Details ───").font = Font(name='Segoe UI', size=10,
    bold=True, italic=True, color=MID_BG)

for ri, (field, sc) in enumerate(zip(fields, src_cols), 6):
    lbl = ws_look.cell(row=ri, column=1, value=field)
    lbl.font   = body_font(10, bold=True)
    lbl.fill   = fill(LIGHT_BG)
    lbl.border = thin_border()
    lbl.alignment = vcenter()

    val_formula = (
        f'=IFERROR(INDEX(Employee_Data!{get_column_letter(sc)}:' +
        f'{get_column_letter(sc)},MATCH($B$3,Employee_Data!$A:$A,0)),"Not Found")'
    )
    vc = ws_look.cell(row=ri, column=2, value=val_formula)
    vc.font   = body_font(10)
    vc.border = thin_border()
    vc.alignment = vcenter()
    vc.fill = fill(WHITE)
    if field == "Monthly Salary":
        vc.number_format = "₹#,##0"

# Data validation for Emp ID dropdown
dv = DataValidation(type="list",
    formula1='"' + ",".join([f"EMP{i:03d}" for i in range(1, 21)]) + '"',
    allow_blank=True,
    showErrorMessage=True)
dv.error   = "Please select a valid Employee ID"
dv.errorTitle = "Invalid ID"
ws_look.add_data_validation(dv)
dv.add("B3")

for col, w in zip("ABCDEFG", [20, 22, 5, 5, 5, 5, 5]):
    ws_look.column_dimensions[col].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 – SALARY ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
ws_sal = wb.create_sheet("Salary_Analysis")

ws_sal.merge_cells("A1:K1")
sb = ws_sal.cell(row=1, column=1, value="💰  SALARY BAND & COMPENSATION ANALYSIS")
sb.font = Font(name='Segoe UI', size=14, bold=True, color=WHITE)
sb.fill = fill("1A6B4A"); sb.alignment = center()
ws_sal.row_dimensions[1].height = 36

section_header(ws_sal, 3, 1, "SALARY BANDS BY DEPARTMENT", span=7)
sal_h = ["Department","Min Salary","Max Salary","Avg Salary",
         "Median*","Total Payroll","% of Total Payroll"]
for i, h in enumerate(sal_h, 1):
    c = ws_sal.cell(row=4, column=i, value=h)
    c.font = header_font(9); c.fill = fill(MID_BG); c.alignment = center()

for ri, dept in enumerate(dept_list, 5):
    ws_sal.cell(row=ri, column=1, value=dept).font = body_font(9, bold=True)
    ws_sal.cell(row=ri, column=2,
        value=f'=IFERROR(MINIFS(Employee_Data!H$2:H$1000,Employee_Data!C$2:C$1000,A{ri}),0)'
        ).number_format="₹#,##0"
    ws_sal.cell(row=ri, column=3,
        value=f'=IFERROR(MAXIFS(Employee_Data!H$2:H$1000,Employee_Data!C$2:C$1000,A{ri}),0)'
        ).number_format="₹#,##0"
    ws_sal.cell(row=ri, column=4,
        value=f'=IFERROR(AVERAGEIF(Employee_Data!C$2:C$1000,A{ri},Employee_Data!H$2:H$1000),0)'
        ).number_format="₹#,##0"
    ws_sal.cell(row=ri, column=5,
        value=f'=IFERROR(AVERAGEIF(Employee_Data!C$2:C$1000,A{ri},Employee_Data!H$2:H$1000),0)'
        ).number_format="₹#,##0"
    ws_sal.cell(row=ri, column=6,
        value=f'=IFERROR(SUMIF(Employee_Data!C$2:C$1000,A{ri},Employee_Data!H$2:H$1000),0)'
        ).number_format="₹#,##0"
    ws_sal.cell(row=ri, column=7,
        value=f'=IFERROR(F{ri}/SUM($F$5:$F$11),0)').number_format="0.0%"
    for c in range(1, 8):
        ws_sal.cell(row=ri, column=c).border    = thin_border()
        ws_sal.cell(row=ri, column=c).alignment = center()
        ws_sal.cell(row=ri, column=c).fill = fill(WHITE if ri%2==0 else LIGHT_BG)
        if c != 1:
            ws_sal.cell(row=ri, column=c).font = body_font(9)

# Total row
tr = 12
for c in range(1, 8):
    cell = ws_sal.cell(row=tr, column=c)
    cell.fill = fill(DARK_BG); cell.font = header_font(9); cell.alignment = center(); cell.border = thin_border()
ws_sal.cell(row=tr, column=1, value="GRAND TOTAL")
ws_sal.cell(row=tr, column=2, value="=MIN(Employee_Data!H2:H1000)").number_format = "₹#,##0"
ws_sal.cell(row=tr, column=3, value="=MAX(Employee_Data!H2:H1000)").number_format = "₹#,##0"
ws_sal.cell(row=tr, column=4, value="=AVERAGE(Employee_Data!H2:H1000)").number_format = "₹#,##0"
ws_sal.cell(row=tr, column=6, value="=SUM(F5:F11)").number_format = "₹#,##0"
ws_sal.cell(row=tr, column=7, value="=SUM(G5:G11)").number_format = "0.0%"

# Salary heatmap note
ws_sal.cell(row=14, column=1,
    value="* Median approximated as Average. Use PERCENTILE() in Excel 365 for exact median by dept.").font = Font(
    name='Segoe UI', size=8, italic=True, color="888888")

# Conditional format – total payroll data bar
ws_sal.conditional_formatting.add(
    "F5:F11",
    DataBarRule(start_type='min', start_value=0,
                end_type='max', end_value=None, color="1A6B4A")
)

sal_widths = [16,14,14,14,14,16,18]
for i, w in enumerate(sal_widths, 1):
    ws_sal.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 – INSTRUCTIONS / COVER PAGE
# ════════════════════════════════════════════════════════════════════════════
ws_cover = wb.create_sheet("📋 Cover", 0)

ws_cover.column_dimensions['A'].width = 4
ws_cover.column_dimensions['B'].width = 36
ws_cover.column_dimensions['C'].width = 54

def cover_row(ws, row, label, value, label_color=DARK_BG, value_color=TEXT_DARK, h=20):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = Font(name='Segoe UI', size=10, bold=True, color=WHITE)
    lc.fill = fill(label_color); lc.alignment = vcenter(); lc.border = thin_border()
    vc = ws.cell(row=row, column=3, value=value)
    vc.font = body_font(10, color=value_color)
    vc.fill = fill(WHITE); vc.alignment = vcenter(); vc.border = thin_border()
    ws.row_dimensions[row].height = h

ws_cover.merge_cells("B1:E3")
title = ws_cover.cell(row=1, column=2,
    value="HR ANALYTICS DASHBOARD\nAdvanced Excel Portfolio Project")
title.font = Font(name='Segoe UI', size=18, bold=True, color=WHITE)
title.fill = fill(DARK_BG); title.alignment = center()
ws_cover.row_dimensions[1].height = 70

cover_row(ws_cover, 5,  "📁 Project Name",    "HR Workforce Analytics Dashboard")
cover_row(ws_cover, 6,  "🎯 Purpose",         "End-to-end HR data analysis using advanced Excel — suitable for Data Analyst resume")
cover_row(ws_cover, 7,  "📊 Data Volume",     "100 Employees · 13 Attributes · 7 Departments · 4 Cities")
cover_row(ws_cover, 8,  "🛠 Key Excel Skills",
    "COUNTIFS · AVERAGEIFS · SUMIFS · MINIFS · MAXIFS · INDEX-MATCH · Conditional Formatting · Data Validation · Charts")
cover_row(ws_cover, 9,  "📑 Sheets",
    "Cover | Employee Data | Summary Dashboard | Charts | Attrition Analysis | Employee Lookup | Salary Analysis")
cover_row(ws_cover, 10, "✅ Features",
    "KPI Cards · Department Breakdown · Performance Split · Attrition Rate · Salary Bands · Dynamic Lookup Tool · 4 Charts")
cover_row(ws_cover, 11, "💡 Resume Bullet",
    'Built an HR Analytics Dashboard in Excel analyzing 100+ employee records across 7 depts using COUNTIFS, INDEX-MATCH, and Conditional Formatting to surface attrition, salary, and performance insights')

ws_cover.merge_cells("B13:E13")
nav = ws_cover.cell(row=13, column=2,
    value="▶  NAVIGATION — Click a sheet tab below to explore each section")
nav.font = Font(name='Segoe UI', size=11, bold=True, color=WHITE)
nav.fill = fill(MID_BG); nav.alignment = center()
ws_cover.row_dimensions[13].height = 28

nav_items = [
    ("📋 Cover",          "This page — project overview"),
    ("Employee_Data",     "Raw data table — 100 employees, 13 fields"),
    ("Summary",           "KPI cards + dept/performance breakdown (formula-powered)"),
    ("Charts",            "4 embedded charts — bar, pie, line"),
    ("Attrition_Analysis","Attrition rates by dept & experience band"),
    ("Employee_Lookup",   "Dynamic employee card using INDEX-MATCH"),
    ("Salary_Analysis",   "Salary bands — min/max/avg/payroll by dept"),
]
for ri, (sheet, desc) in enumerate(nav_items, 14):
    ws_cover.cell(row=ri, column=2, value=f"→  {sheet}").font = body_font(9, bold=True, color=MID_BG)
    ws_cover.cell(row=ri, column=2).fill = fill(LIGHT_BG)
    ws_cover.cell(row=ri, column=2).border = thin_border()
    ws_cover.cell(row=ri, column=2).alignment = vcenter()
    ws_cover.cell(row=ri, column=3, value=desc).font = body_font(9)
    ws_cover.cell(row=ri, column=3).fill = fill(WHITE)
    ws_cover.cell(row=ri, column=3).border = thin_border()
    ws_cover.cell(row=ri, column=3).alignment = vcenter()
    ws_cover.row_dimensions[ri].height = 20

# Tab colours
ws_cover.sheet_properties.tabColor   = DARK_BG
ws_data.sheet_properties.tabColor    = MID_BG
ws_sum.sheet_properties.tabColor     = ACCENT
ws_chart.sheet_properties.tabColor   = GOLD
ws_attr.sheet_properties.tabColor    = RED
ws_look.sheet_properties.tabColor    = GREEN
ws_sal.sheet_properties.tabColor     = "1A6B4A"

# Save
out = "/mnt/user-data/outputs/HR_Analytics_Dashboard.xlsx"
wb.save(out)
print("Saved:", out)
